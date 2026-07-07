#!/usr/bin/env python3
"""
Find all tables in Synapse (dev + prod) that do NOT have an ACE_TIMESTAMP column,
and list all Databricks notebooks (dev + prod) with their owner/creator.
Outputs a multi-sheet Excel file (with CSV fallback if openpyxl is unavailable).
"""

import csv
import struct
import subprocess
from datetime import datetime, timezone

import pyodbc
import requests

# ── Synapse environments ───────────────────────────────────────────────────────

SYNAPSE_ENVS = [
    {
        "label": "DEV",
        "server":   "zus1-idoh-dev-v2-sql-server.database.windows.net",
        "database": "zus1-idoh-dev-v2-sql-dw",
    },
    {
        "label": "PRD",
        "server":   "zus1-idoh-prd-v1-sql-server.database.windows.net",
        "database": "zus1-idoh-prd-v1-sql-dw",
    },
]

# ── Databricks environments ────────────────────────────────────────────────────

DATABRICKS_ENVS = [
    {
        "label":    "DEV",
        "hostname": "adb-5757046586469840.0.azuredatabricks.net",
        "token":    os.environ.get("DATABRICKS_DEV_TOKEN", ""),  # set env var or replace with your token
    },
    {
        "label":    "PRD",
        "hostname": "adb-5323951998838804.4.azuredatabricks.net",
        "token":    os.environ.get("DATABRICKS_PRD_TOKEN", ""),  # set env var or replace with your token
    },
]

DRIVER      = "{ODBC Driver 18 for SQL Server}"
OUTPUT_XLSX = "/home/thedavidporter/synapse_missing_ace_timestamp.xlsx"
OUTPUT_CSV  = "/home/thedavidporter/synapse_missing_ace_timestamp.csv"

# ── Synapse auth + query ───────────────────────────────────────────────────────

def get_synapse_connection(server, database):
    result = subprocess.run(
        ["az", "account", "get-access-token",
         "--resource", "https://database.windows.net",
         "--query", "accessToken", "-o", "tsv"],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"az login required: {result.stderr.strip()}")
    token        = result.stdout.strip()
    token_bytes  = token.encode("utf-16-le")
    token_struct = struct.pack(f"<I{len(token_bytes)}s", len(token_bytes), token_bytes)
    conn_str = (
        f"Driver={DRIVER};Server=tcp:{server},1433;"
        f"Database={database};Encrypt=yes;TrustServerCertificate=no;"
    )
    return pyodbc.connect(conn_str, attrs_before={1256: token_struct})

def sql_query(conn, sql):
    cur = conn.cursor()
    cur.execute(sql)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

# ── Synapse SQL ────────────────────────────────────────────────────────────────

MISSING_ACE_SQL = """
SELECT
    t.TABLE_SCHEMA  AS schema_name,
    t.TABLE_NAME    AS table_name
FROM INFORMATION_SCHEMA.TABLES t
WHERE t.TABLE_TYPE = 'BASE TABLE'
  AND NOT EXISTS (
      SELECT 1
      FROM INFORMATION_SCHEMA.COLUMNS c
      WHERE c.TABLE_SCHEMA = t.TABLE_SCHEMA
        AND c.TABLE_NAME   = t.TABLE_NAME
        AND UPPER(c.COLUMN_NAME) = 'ACE_TIMESTAMP'
  )
ORDER BY t.TABLE_SCHEMA, t.TABLE_NAME
"""

ROW_COUNTS_SQL = """
SELECT s.name AS schema_name, t.name AS table_name, SUM(nps.row_count) AS row_count
FROM sys.tables t
JOIN sys.schemas s ON t.schema_id = s.schema_id
JOIN sys.pdw_table_mappings tm ON tm.object_id = t.object_id
JOIN sys.pdw_nodes_tables nt ON nt.name = tm.physical_name
JOIN sys.dm_pdw_nodes_db_partition_stats nps
    ON nps.object_id        = nt.object_id
    AND nps.pdw_node_id     = nt.pdw_node_id
    AND nps.distribution_id = nt.distribution_id
    AND nps.index_id < 2
GROUP BY s.name, t.name
"""

# ── Synapse data collection ────────────────────────────────────────────────────

def collect_synapse(env):
    label    = env["label"]
    server   = env["server"]
    database = env["database"]

    print(f"\n[Synapse {label}] Connecting to {database}…")
    conn = get_synapse_connection(server, database)
    print(f"[Synapse {label}] Connected.")

    print(f"[Synapse {label}]   Querying tables missing ACE_TIMESTAMP…", end="", flush=True)
    try:
        missing = sql_query(conn, MISSING_ACE_SQL)
        print(f" {len(missing)} tables")
    except Exception as e:
        print(f" ERROR: {e}")
        missing = []

    print(f"[Synapse {label}]   Querying row counts…", end="", flush=True)
    try:
        rc_rows = sql_query(conn, ROW_COUNTS_SQL)
        print(f" {len(rc_rows)} rows")
    except Exception as e:
        print(f" ERROR: {e}")
        rc_rows = []

    conn.close()

    rc_map = {(r["schema_name"], r["table_name"]): r["row_count"] for r in rc_rows}

    rows = []
    for t in missing:
        s, n = t["schema_name"], t["table_name"]
        rc = rc_map.get((s, n), None)
        rows.append({
            "Environment": label,
            "Database":    database,
            "Schema":      s,
            "Table":       n,
            "Row Count":   int(rc) if rc is not None else "",
        })
    return rows

# ── Databricks REST API ────────────────────────────────────────────────────────

def _db_list(hostname, token, path):
    """Call workspace/list for a single path; returns list of ObjectInfo dicts."""
    url  = f"https://{hostname}/api/2.0/workspace/list"
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params={"path": path},
        timeout=30,
    )
    if resp.status_code == 404:
        return []
    resp.raise_for_status()
    return resp.json().get("objects", [])

def _db_get_status(hostname, token, path):
    """Call workspace/get-status for a single notebook path to fetch creator_user_name."""
    url  = f"https://{hostname}/api/2.0/workspace/get-status"
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params={"path": path},
        timeout=30,
    )
    if not resp.ok:
        return {}
    return resp.json()

def _fmt_ts(ms):
    """Convert Databricks millisecond epoch to a readable date string."""
    if not ms:
        return ""
    try:
        return datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        return ""

def collect_databricks_notebooks(env):
    label    = env["label"]
    hostname = env["hostname"]
    token    = env["token"]

    print(f"\n[Databricks {label}] Walking workspace {hostname}…")

    notebooks = []
    stack = ["/"]
    visited = set()

    while stack:
        path = stack.pop()
        if path in visited:
            continue
        visited.add(path)

        try:
            objects = _db_list(hostname, token, path)
        except Exception as e:
            print(f"  WARN: could not list {path}: {e}")
            continue

        for obj in objects:
            obj_type = obj.get("object_type", "")
            obj_path = obj.get("path", "")

            if obj_type == "DIRECTORY":
                stack.append(obj_path)
            elif obj_type == "NOTEBOOK":
                owner = obj.get("creator_user_name", "")

                # If owner not in list response, call get-status (fallback)
                if not owner:
                    status = _db_get_status(hostname, token, obj_path)
                    owner  = status.get("creator_user_name", "")

                notebooks.append({
                    "Environment":   label,
                    "Workspace":     hostname,
                    "Notebook Path": obj_path,
                    "Language":      obj.get("language", ""),
                    "Owner":         owner,
                    "Last Modified": _fmt_ts(obj.get("modified_at")),
                })

    print(f"[Databricks {label}] Found {len(notebooks)} notebooks.")
    return notebooks

# ── Excel output ───────────────────────────────────────────────────────────────

SYNAPSE_COLS   = ["Environment", "Database", "Schema", "Table", "Row Count"]
NOTEBOOK_COLS  = ["Environment", "Workspace", "Notebook Path", "Language", "Owner", "Last Modified"]

def _style_header_row(ws, cols, title_text, title_bg, header_bg, get_col_letter):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    n = len(cols)
    last_col = get_col_letter(n)
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value     = title_text
    tc.font      = Font(bold=True, size=12, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=title_bg)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    hdr_fill   = PatternFill("solid", fgColor=header_bg)
    hdr_font   = Font(bold=True, color="FFFFFF")
    thin_right = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="CCCCCC"),
    )
    for ci, name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_right

def _write_data_rows(ws, rows, cols, env_fills, alt_fills):
    from openpyxl.styles import PatternFill, Alignment, Border, Side

    row_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right =Side(style="thin", color="DDDDDD"),
    )
    for row_idx, row in enumerate(rows, start=3):
        env  = row.get("Environment", "")
        fill = env_fills.get(env) if row_idx % 2 == 1 else alt_fills.get(env)
        for ci, col_name in enumerate(cols, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = row_border
            if col_name == "Row Count" and isinstance(val, int):
                cell.number_format = "#,##0"
                cell.alignment     = Alignment(horizontal="right")

def _autofit(ws, rows, cols, get_col_letter):
    widths = {c: len(c) + 2 for c in cols}
    for row in rows:
        for c in cols:
            v = str(row.get(c, "")) if row.get(c, "") != "" else ""
            widths[c] = max(widths[c], len(v) + 2)
    for ci, c in enumerate(cols, start=1):
        ws.column_dimensions[get_col_letter(ci)].width = min(widths[c], 60)

def write_excel(synapse_rows, notebook_rows, generated):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from collections import Counter

    wb = openpyxl.Workbook()

    env_fills = {
        "DEV": PatternFill("solid", fgColor="EBF3FB"),
        "PRD": PatternFill("solid", fgColor="FEF9E7"),
    }
    alt_fills = {
        "DEV": PatternFill("solid", fgColor="D6EAF8"),
        "PRD": PatternFill("solid", fgColor="FDF2E9"),
    }

    # ── Sheet 1: Synapse tables missing ACE_TIMESTAMP ──────────────────────────
    ws1 = wb.active
    ws1.title = "Missing ACE_TIMESTAMP"
    _style_header_row(
        ws1, SYNAPSE_COLS,
        f"Synapse Tables Missing ACE_TIMESTAMP — Generated {generated}",
        "1F3864", "2E4D7B", get_column_letter,
    )
    _write_data_rows(ws1, synapse_rows, SYNAPSE_COLS, env_fills, alt_fills)
    _autofit(ws1, synapse_rows, SYNAPSE_COLS, get_column_letter)
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Databricks notebooks with owner ───────────────────────────────
    ws2 = wb.create_sheet("Databricks Notebooks")
    _style_header_row(
        ws2, NOTEBOOK_COLS,
        f"Databricks Notebook Owners — Generated {generated}",
        "1A3A1A", "2E6B2E", get_column_letter,
    )
    _write_data_rows(ws2, notebook_rows, NOTEBOOK_COLS, env_fills, alt_fills)
    _autofit(ws2, notebook_rows, NOTEBOOK_COLS, get_column_letter)
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Summary ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    headers = ["Category", "Environment", "Detail", "Count"]
    for ci, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)

    summary_rows = []

    # Synapse counts
    for (env, db), cnt in sorted(Counter((r["Environment"], r["Database"]) for r in synapse_rows).items()):
        summary_rows.append(["Synapse – Tables Missing ACE_TIMESTAMP", env, db, cnt])

    # Databricks counts by environment
    for (env, ws_host), cnt in sorted(Counter((r["Environment"], r["Workspace"]) for r in notebook_rows).items()):
        summary_rows.append(["Databricks – Total Notebooks", env, ws_host, cnt])

    # Databricks notebook counts by owner (top owners)
    owner_counts = Counter(r["Owner"] for r in notebook_rows if r["Owner"])
    for owner, cnt in owner_counts.most_common():
        summary_rows.append(["Databricks – Notebooks by Owner", "", owner, cnt])

    for ri, row in enumerate(summary_rows, start=2):
        for ci, val in enumerate(row, start=1):
            ws3.cell(row=ri, column=ci, value=val)

    for col in ["A", "B", "C", "D"]:
        ws3.column_dimensions[col].width = 50 if col == "C" else 35

    wb.save(OUTPUT_XLSX)
    print(f"\nExcel saved to: {OUTPUT_XLSX}")

# ── CSV fallback ───────────────────────────────────────────────────────────────

def write_csv(synapse_rows, notebook_rows):
    synapse_csv  = OUTPUT_CSV.replace(".csv", "_synapse.csv")
    notebook_csv = OUTPUT_CSV.replace(".csv", "_databricks.csv")

    with open(synapse_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SYNAPSE_COLS)
        writer.writeheader()
        writer.writerows(synapse_rows)
    print(f"CSV saved to: {synapse_csv}")

    with open(notebook_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=NOTEBOOK_COLS)
        writer.writeheader()
        writer.writerows(notebook_rows)
    print(f"CSV saved to: {notebook_csv}")

# ── entry point ────────────────────────────────────────────────────────────────

def main():
    generated     = datetime.now().strftime("%Y-%m-%d %H:%M")
    synapse_rows  = []
    notebook_rows = []

    # Synapse: tables missing ACE_TIMESTAMP
    for env in SYNAPSE_ENVS:
        try:
            synapse_rows.extend(collect_synapse(env))
        except Exception as e:
            print(f"\n[Synapse {env['label']}] FAILED: {e}")

    # Databricks: notebook owners
    for env in DATABRICKS_ENVS:
        try:
            notebook_rows.extend(collect_databricks_notebooks(env))
        except Exception as e:
            print(f"\n[Databricks {env['label']}] FAILED: {e}")

    # Summary to console
    print(f"\n{'='*60}")
    print(f"Synapse tables missing ACE_TIMESTAMP : {len(synapse_rows)}")
    from collections import Counter
    for (env, db), cnt in sorted(Counter((r["Environment"], r["Database"]) for r in synapse_rows).items()):
        print(f"  {env} ({db}): {cnt}")

    print(f"\nDatabricks notebooks found           : {len(notebook_rows)}")
    for (env, ws), cnt in sorted(Counter((r["Environment"], r["Workspace"]) for r in notebook_rows).items()):
        print(f"  {env} ({ws}): {cnt}")

    if not synapse_rows and not notebook_rows:
        print("\nNo data collected — check your credentials and network access.")
        return

    try:
        write_excel(synapse_rows, notebook_rows, generated)
    except ImportError:
        print("\nopenpyxl not installed — writing CSV instead.")
        write_csv(synapse_rows, notebook_rows)

if __name__ == "__main__":
    main()
