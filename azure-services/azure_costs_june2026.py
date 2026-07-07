"""
azure_costs_june2026.py  —  Build Azure cost Excel from a Cost Management export CSV.

Usage:
    python3 azure_costs_june2026.py <path/to/export.csv> [--start YYYY-MM-DD] [--end YYYY-MM-DD] [--out output.xlsx]

Download the CSV from:
    Azure Portal → Cost Management + Billing → Cost Analysis → Download (CSV)
    or: Subscriptions → <sub> → Cost analysis → Download
"""

import sys
import argparse
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Column detection ───────────────────────────────────────────────────────────

# Priority-ordered candidates for each logical field.
SUBSCRIPTION_CANDIDATES = ["subscriptionname", "subscription"]
SERVICE_CANDIDATES      = ["servicename", "metercategory", "consumedservice"]
COST_CANDIDATES         = ["pretaxcost", "costinbillingcurrency", "cost", "extendedcost", "effectivecost"]
DATE_CANDIDATES         = ["date", "usagedatetime", "usagedate", "billingperiodstartdate"]


def _pick(col_map, candidates, label):
    for c in candidates:
        if c in col_map:
            return col_map[c]
    raise SystemExit(
        f"ERROR: Could not find a '{label}' column.\n"
        f"  Looked for: {candidates}\n"
        f"  Found in CSV: {sorted(col_map.keys())}"
    )


def detect_columns(df):
    col_map = {c.strip().lower(): c for c in df.columns}
    sub_col  = _pick(col_map, SUBSCRIPTION_CANDIDATES, "SubscriptionName")
    svc_col  = _pick(col_map, SERVICE_CANDIDATES,      "ServiceName")
    cost_col = _pick(col_map, COST_CANDIDATES,         "Cost")
    date_col = _pick(col_map, DATE_CANDIDATES,         "Date")
    return sub_col, svc_col, cost_col, date_col


# ── Data loading ───────────────────────────────────────────────────────────────

def load_csv(path, start=None, end=None):
    # Azure exports sometimes have a metadata preamble before the header row.
    # Skip rows until we find one containing "SubscriptionName" or "Subscription".
    skip = 0
    with open(path, encoding="utf-8-sig") as f:
        for i, line in enumerate(f):
            lower = line.lower()
            if any(c in lower for c in SUBSCRIPTION_CANDIDATES + SERVICE_CANDIDATES):
                skip = i
                break

    df = pd.read_csv(path, skiprows=skip, low_memory=False)
    df.columns = df.columns.str.strip()

    sub_col, svc_col, cost_col, date_col = detect_columns(df)

    df[cost_col] = pd.to_numeric(df[cost_col], errors="coerce").fillna(0)
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

    if start:
        df = df[df[date_col] >= pd.Timestamp(start)]
    if end:
        df = df[df[date_col] <= pd.Timestamp(end)]

    return df, sub_col, svc_col, cost_col, date_col


def build_data(df, sub_col, svc_col, cost_col):
    df = df[df[cost_col] > 0].copy()
    grouped = (
        df.groupby([sub_col, svc_col], dropna=False)[cost_col]
        .sum()
        .reset_index()
    )
    data = defaultdict(list)
    for _, row in grouped.iterrows():
        sub  = str(row[sub_col]).strip()
        svc  = str(row[svc_col]).strip()
        cost = float(row[cost_col])
        data[sub].append((svc, cost))
    return dict(data)


# ── Excel generation ───────────────────────────────────────────────────────────

HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")
SUB_HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")
TOTAL_FILL       = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL         = PatternFill("solid", fgColor="EBF3FB")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="1F4E79")
WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")

_thin   = Side(style="thin", color="BFBFBF")
BORDER  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_row(ws, values, fill=SUB_HEADER_FILL):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER


def _data_row(ws, values, i, cost_col_idx=2):
    ws.append(values)
    r    = ws.max_row
    fill = ALT_FILL if i % 2 else WHITE_FILL
    for cell in ws[r]:
        cell.border = BORDER
        cell.fill   = fill
    ws.cell(r, cost_col_idx).number_format  = '"$"#,##0.00'
    ws.cell(r, cost_col_idx).alignment      = Alignment(horizontal="right")


def _total_row(ws, label, total, pct_label="100.0%"):
    ws.append([label, round(total, 2), pct_label])
    r = ws.max_row
    for cell in ws[r]:
        cell.font   = Font(bold=True, color="FFFFFF")
        cell.fill   = GRAND_TOTAL_FILL
        cell.border = BORDER
    ws.cell(r, 2).number_format = '"$"#,##0.00'
    ws.cell(r, 2).alignment     = Alignment(horizontal="right")
    ws.cell(r, 3).alignment     = Alignment(horizontal="center")


def _set_widths(ws, widths):
    for col, w in zip("ABCD", widths):
        ws.column_dimensions[col].width = w


def build_excel(data, date_range_str, out_path):
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _set_widths(ws, [38, 18, 14])

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = f"Azure Cost Summary — {date_range_str}"
    t.font      = Font(bold=True, size=14, color="FFFFFF")
    t.fill      = HEADER_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    _hdr_row(ws, ["Subscription", "MTD Cost (USD)", "% of Total"])

    sub_totals  = {sub: sum(c for _, c in rows) for sub, rows in data.items()}
    grand_total = sum(sub_totals.values())

    for i, (sub, total) in enumerate(sorted(sub_totals.items(), key=lambda x: -x[1])):
        pct = total / grand_total * 100 if grand_total else 0
        _data_row(ws, [sub, round(total, 2), f"{pct:.1f}%"], i)
        ws.cell(ws.max_row, 3).alignment = Alignment(horizontal="center")

    _total_row(ws, "TOTAL", grand_total)

    # ── Per-subscription sheets ────────────────────────────────────
    for sub, rows in sorted(data.items(), key=lambda x: -sum(c for _, c in x[1])):
        ws2 = wb.create_sheet(title=sub[:31])
        _set_widths(ws2, [38, 18, 14])

        ws2.merge_cells("A1:C1")
        t2            = ws2["A1"]
        t2.value      = f"{sub} — {date_range_str}"
        t2.font       = Font(bold=True, size=13, color="FFFFFF")
        t2.fill       = HEADER_FILL
        t2.alignment  = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26

        _hdr_row(ws2, ["Service", "Cost (USD)", "% of Sub Total"])

        sub_total = sum(c for _, c in rows)
        for i, (svc, cost) in enumerate(sorted(rows, key=lambda x: -x[1])):
            pct = cost / sub_total * 100 if sub_total else 0
            _data_row(ws2, [svc, round(cost, 4), f"{pct:.1f}%"], i)
            ws2.cell(ws2.max_row, 3).alignment = Alignment(horizontal="center")

        _total_row(ws2, "TOTAL", sub_total)

    wb.save(out_path)
    print(f"Saved: {out_path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build Azure cost Excel from Cost Management CSV export.")
    parser.add_argument("csv", help="Path to Azure Cost Management CSV export")
    parser.add_argument("--start", metavar="YYYY-MM-DD", help="Include rows on or after this date")
    parser.add_argument("--end",   metavar="YYYY-MM-DD", help="Include rows on or before this date")
    parser.add_argument("--out",   default="/home/thedavidporter/azure_costs_june2026.xlsx",
                        help="Output .xlsx path (default: azure_costs_june2026.xlsx)")
    args = parser.parse_args()

    df, sub_col, svc_col, cost_col, date_col = load_csv(args.csv, args.start, args.end)

    # Build a human-readable date range string for titles
    dates = df[date_col].dropna()
    if not dates.empty:
        lo = dates.min().strftime("%-m/%-d/%Y")
        hi = dates.max().strftime("%-m/%-d/%Y")
        date_range_str = lo if lo == hi else f"{lo} – {hi}"
    else:
        date_range_str = "Date range unknown"

    print(f"Loaded {len(df):,} rows  |  {df[sub_col].nunique()} subscriptions  |  "
          f"date range: {date_range_str}")

    data = build_data(df, sub_col, svc_col, cost_col)

    print(f"Subscriptions found: {sorted(data.keys())}")

    build_excel(data, date_range_str, args.out)


if __name__ == "__main__":
    main()
