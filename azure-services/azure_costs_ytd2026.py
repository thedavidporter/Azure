import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = {
    "ECAE IDOH Production": [
        ("Azure Databricks", 11673.02),
        ("Virtual Machines", 7228.66),
        ("Storage", 3014.91),
        ("Azure Data Factory v2", 2798.00),
        ("Azure DevOps", 2144.56),
        ("Microsoft Defender for Cloud", 1267.29),
        ("Log Analytics", 790.50),
        ("Azure Synapse Analytics", 713.13),
        ("Container Registry", 564.88),
        ("Virtual Network", 408.17),
        ("Logic Apps", 46.55),
        ("Azure Monitor", 29.97),
        ("Container Instances", 14.68),
        ("Azure DNS", 7.72),
        ("Bandwidth", 2.91),
        ("Event Grid", 1.62),
        ("Key Vault", 0.32),
        ("NAT Gateway", 0.01),
    ],
    "ECAE DWD Production": [
        ("SQL Database", 28308.06),
        ("Log Analytics", 19001.64),
        ("Azure Synapse Analytics", 12552.86),
        ("Storage", 11483.05),
        ("Virtual Machines", 7344.17),
        ("Azure Databricks", 7275.28),
        ("Microsoft Defender for Cloud", 3954.40),
        ("Logic Apps", 976.57),
        ("Virtual Network", 424.40),
        ("Azure DevOps", 392.03),
        ("Azure Data Factory v2", 211.27),
        ("Event Grid", 122.30),
        ("Azure Monitor", 38.29),
        ("Azure DNS", 10.54),
        ("Bandwidth", 6.56),
        ("Key Vault", 0.31),
        ("NAT Gateway", 0.04),
        ("Azure Cosmos DB", 0.0),
        ("Azure App Service", 0.0),
    ],
    "ECAE DOE Production": [
        ("Azure Synapse Analytics", 34102.08),
        ("Storage", 2063.49),
        ("Microsoft Defender for Cloud", 1633.07),
        ("Virtual Machines", 811.52),
        ("Log Analytics", 901.97),
        ("Azure Data Factory v2", 791.61),
        ("Azure Databricks", 71.91),
        ("Virtual Network", 45.72),
        ("Azure Monitor", 0.46),
        ("Bandwidth", 2.60),
        ("Event Grid", 0.21),
        ("Key Vault", 0.06),
    ],
    "ECAE Shared Production": [
        ("Virtual Machines", 42015.03),
        ("Storage", 11935.99),
        ("Azure Firewall", 7812.79),
        ("Microsoft Defender for Cloud", 7728.29),
        ("Log Analytics", 5941.55),
        ("Azure Database for PostgreSQL", 3062.10),
        ("Application Gateway", 2050.76),
        ("Virtual Network", 1717.38),
        ("Azure DevOps", 436.27),
        ("Azure Kubernetes Service", 406.83),
        ("Container Registry", 282.44),
        ("Load Balancer", 203.63),
        ("Bandwidth", 82.07),
        ("Key Vault", 37.05),
        ("Automation", 22.82),
        ("Azure DNS", 5.40),
        ("Azure Monitor", 0.52),
        ("Event Grid", 0.0),
    ],
    "ECAE Shared Dev": [
        ("Virtual Machines", 7537.32),
        ("Azure Firewall", 7423.68),
        ("Storage", 6230.42),
        ("Microsoft Defender for Cloud", 2313.50),
        ("Azure Database for PostgreSQL", 2294.99),
        ("Virtual Network", 2168.22),
        ("Application Gateway", 2050.76),
        ("Log Analytics", 932.23),
        ("Azure Kubernetes Service", 649.27),
        ("Container Registry", 282.44),
        ("Load Balancer", 264.34),
        ("Key Vault", 37.06),
        ("Bandwidth", 32.00),
        ("Azure DNS", 9.46),
        ("Automation", 4.08),
        ("Event Grid", 0.0),
    ],
    "ISDH Production": [
        ("SQL Database", 20994.61),
        ("Microsoft Defender for Cloud", 1970.59),
        ("Virtual Machines", 1854.14),
        ("Azure App Service", 1610.79),
        ("Azure DevOps", 1461.45),
        ("Log Analytics", 1307.02),
        ("GitHub", 1265.94),
        ("Storage", 791.10),
        ("Virtual Network", 760.22),
        ("Container Registry", 564.88),
        ("Azure Container Apps", 538.20),
        ("VPN Gateway", 292.97),
        ("Load Balancer", 203.45),
        ("Azure Data Factory v2", 97.28),
        ("Foundry Tools", 54.78),
        ("Azure DNS", 29.71),
        ("Bandwidth", 17.32),
        ("Key Vault", 8.92),
        ("Azure Monitor", 0.91),
        ("Event Grid", 0.02),
        ("Logic Apps", 0.0),
        ("Azure Cognitive Search", 0.0),
    ],
    "ECAE IOT Production": [
        ("Microsoft Defender for Cloud", 165.73),
        ("Storage", 1.38),
        ("Log Analytics", 1.01),
        ("Event Grid", 0.01),
        ("Bandwidth", 0.0),
    ],
}

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="1F4E79")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Summary sheet ──────────────────────────────────────────────
ws_summary = wb.active
ws_summary.title = "Summary"

ws_summary.column_dimensions["A"].width = 38
ws_summary.column_dimensions["B"].width = 18
ws_summary.column_dimensions["C"].width = 14

# Title
ws_summary.merge_cells("A1:C1")
title_cell = ws_summary["A1"]
title_cell.value = "Azure Cost Summary — Jan 1 – Jun 10, 2026"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = HEADER_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 28

ws_summary.append(["Subscription", "YTD Cost (USD)", "% of Total"])
for cell in ws_summary[2]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = SUB_HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

sub_totals = {sub: sum(c for _, c in rows) for sub, rows in data.items()}
grand_total = sum(sub_totals.values())

for i, (sub, total) in enumerate(sorted(sub_totals.items(), key=lambda x: -x[1])):
    pct = total / grand_total * 100
    row = [sub, round(total, 2), f"{pct:.1f}%"]
    ws_summary.append(row)
    r = ws_summary.max_row
    fill = ALT_FILL if i % 2 else PatternFill("solid", fgColor="FFFFFF")
    for cell in ws_summary[r]:
        cell.border = border
        cell.fill = fill
    ws_summary[f"B{r}"].number_format = '"$"#,##0.00'
    ws_summary[f"B{r}"].alignment = Alignment(horizontal="right")
    ws_summary[f"C{r}"].alignment = Alignment(horizontal="center")

# Grand total row
ws_summary.append(["TOTAL (6 of 7 subscriptions, Shared split prod/dev)", round(grand_total, 2), "100.0%"])
r = ws_summary.max_row
for cell in ws_summary[r]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = GRAND_TOTAL_FILL
    cell.border = border
ws_summary[f"B{r}"].number_format = '"$"#,##0.00'
ws_summary[f"B{r}"].alignment = Alignment(horizontal="right")
ws_summary[f"C{r}"].alignment = Alignment(horizontal="center")

ws_summary.append([])
ws_summary.append(["* MPH Production excluded — insufficient permissions"])
ws_summary[ws_summary.max_row][0].font = Font(italic=True, color="808080")

# ── Per-subscription sheets ────────────────────────────────────
for sub, rows in sorted(data.items(), key=lambda x: -sum(c for _, c in x[1])):
    ws = wb.create_sheet(title=sub[:31])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"{sub} — Jan 1 – Jun 10, 2026"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = HEADER_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(["Service", "YTD Cost (USD)", "% of Sub Total"])
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = SUB_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    sub_total = sum(c for _, c in rows)
    for i, (svc, cost) in enumerate(sorted(rows, key=lambda x: -x[1])):
        pct = cost / sub_total * 100 if sub_total else 0
        ws.append([svc, round(cost, 4), f"{pct:.1f}%"])
        r = ws.max_row
        fill = ALT_FILL if i % 2 else PatternFill("solid", fgColor="FFFFFF")
        for cell in ws[r]:
            cell.border = border
            cell.fill = fill
        ws[f"B{r}"].number_format = '"$"#,##0.00'
        ws[f"B{r}"].alignment = Alignment(horizontal="right")
        ws[f"C{r}"].alignment = Alignment(horizontal="center")

    ws.append(["TOTAL", round(sub_total, 2), "100.0%"])
    r = ws.max_row
    for cell in ws[r]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = GRAND_TOTAL_FILL
        cell.border = border
    ws[f"B{r}"].number_format = '"$"#,##0.00'
    ws[f"B{r}"].alignment = Alignment(horizontal="right")
    ws[f"C{r}"].alignment = Alignment(horizontal="center")

output_path = "/home/thedavidporter/azure_costs_ytd2026.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
